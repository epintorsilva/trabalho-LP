import sqlite3
import openpyxl
from flask import render_template
import pyautogui
import time



def entrar_na_prefeitura(senha, cpf, cnpj):
    pyautogui.PAUSE = 1.5

    pyautogui.hotkey('ctrl','t')
    pyautogui.write("https://nfs-e.mage.rj.gov.br")
    pyautogui.press("enter")
    time.sleep(5)
    for i in range(2):
        pyautogui.press('tab')
    pyautogui.press("enter")
    time.sleep(5)

    #preencher os campos
    pyautogui.click(0,200)
    pyautogui.press('tab')
    pyautogui.press('right')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.write(cnpj) #CNPJ do contribuinte
    pyautogui.press('tab')
    pyautogui.write(cpf)# CPF do usuário
    pyautogui.press('tab')
    pyautogui.write(senha) #senha
    pyautogui.press('tab')
    time.sleep(5)
    pyautogui.press('enter')
    time.sleep(5)

def prencher_primeiro_campo(CPF,discriminação_dos_Serviços):
    
    pyautogui.PAUSE = 1.5
    
    #preenchendo NFe
    pyautogui.click(412,549)
    pyautogui.click(482,592)
    time.sleep(2)
    pyautogui.click(482,592)
    time.sleep(5)  
    for i in range(3):
        pyautogui.press('tab')
    pyautogui.write(CPF)
    pyautogui.press('tab')
    pyautogui.press('enter')            
                
    for i in range(6):
        pyautogui.press('tab')
    pyautogui.press('enter')
    for i in range(2):
        pyautogui.press('tab')
    for i in range(2):
        pyautogui.press('down')
    pyautogui.press('enter')

    for i in range(3):
        pyautogui.press('tab')
    pyautogui.write(discriminação_dos_Serviços)
    pyautogui.click(975,585)
    time.sleep(2)

def prencher_segundo_campo(totalServico,faturamento,alicotaISS,valor_licota,PIS,CONFIRNS,IR,INSS,CSLL,retencoes,msg):
    e = faturamento
    g = valor_licota

    pyautogui.PAUSE = 1.5

    pyautogui.write(totalServico)
    pyautogui.press('tab')
    if e == 'não':
        pyautogui.write(alicotaISS)
        pyautogui.press('tab')
        pyautogui.press('tab')
    elif e == 'sim':
        pyautogui.press('tab')
        pyautogui.press('down')
        pyautogui.press('tab')

    if g == 'valor':
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.write(PIS)
        pyautogui.press('tab')
        pyautogui.write(CONFIRNS)
        pyautogui.press('tab')
        pyautogui.write(IR)
        pyautogui.press('tab')
        pyautogui.write(INSS)
        pyautogui.press('tab')
        pyautogui.write(CSLL)
        pyautogui.press('tab')
        pyautogui.write(retencoes)
        pyautogui.press('tab')

    elif g == 'alíquota' or 'aligota':
        pyautogui.press('down')
        pyautogui.press('tab')
        pyautogui.write(PIS)
        pyautogui.press('tab')
        pyautogui.write(CONFIRNS)
        pyautogui.press('tab')
        pyautogui.write(IR)
        pyautogui.press('tab')
        pyautogui.write(INSS)
        pyautogui.press('tab')
        pyautogui.write(CSLL)
        pyautogui.press('tab')
        pyautogui.write(retencoes)
        pyautogui.press('tab')

    pyautogui.write(msg)
    #pyautogui.hotkey('ctrl','s')

def extrair_dados(arquivo):
    book = openpyxl.load_workbook(arquivo)
    page = book.sheetnames[0]
    page = book[page]
    for rows in page.iter_rows(min_row=2):
        
        b = (f'{rows[1].value}')
        c = (f'{rows[2].value}')
        d = (f'{rows[3].value}')
        e = (f'{rows[4].value}')
        f = (f'{rows[5].value}')
        g = (f'{rows[6].value}')
        h = (f'{rows[7].value}')
        i = (f'{rows[8].value}')
        j = (f'{rows[9].value}')
        k = (f'{rows[10].value}')
        l = (f'{rows[11].value}')
        m = (f'{rows[12].value}')
        n = (f'{rows[13].value}')
        if b == 'None':
            break
        else:
            prencher_primeiro_campo(b,c)
            prencher_segundo_campo(d,e,f,g,h,i,j,k,l,m,n)



def auto(arquivo, senha, cpf, cnpj):
    entrar_na_prefeitura(senha, cpf, cnpj)
    extrair_dados(arquivo)
    
