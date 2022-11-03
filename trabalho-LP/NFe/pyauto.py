import sqlite3
import openpyxl
from flask import render_template
import pyautogui
import time


'''
    Este módulo contém funções para automação de emissão de notas fiscais com o autopy, uma
    biblioteca de manipulação do teclado e mouse.
'''

'''
controla teclado e mouse para navegar no siter da NFe da prefeitura de magé
e prenche os campos para efetuar o login.
possui um tempo de espera, de 5s apos preencher todos os campos, para apertar o botao e efetuar o login
isso da ao usuario um tempo para selecionar o recaptcha, o que muitas das vezes não é necessario.
'''
def entrar_na_prefeitura(senha, cpf, cnpj):

    pyautogui.PAUSE = 1.5 #define a velocidade do intervalo de tempo entre uma tecla e outra. note que o 'sleep' mais a frente serve pra prologar esse tempo sem que aumente de todo o codigo

    pyautogui.hotkey('ctrl','t') #abre uma nova guia no browser
    pyautogui.write("https://nfs-e.mage.rj.gov.br") #digita o link do site 
    pyautogui.press("enter")
    time.sleep(5)#time.sleep sao usados para que as chances de erro no codigo por demora de resposta do site seja diminuida
    for i in range(2):#laço de repetição para presionar a tecla 'tab' duas vezes
        pyautogui.press('tab')
    pyautogui.press("enter")
    time.sleep(5)

    #preencher os campos
    pyautogui.click(0,200)#usado somente para clicar no canto da tela e possibilitar que o programa intereja com as teclas a seguir
    pyautogui.press('tab')
    pyautogui.press('right')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.write(cnpj) #CNPJ do contribuinte
    pyautogui.press('tab')
    pyautogui.write(cpf)# CPF do usuário
    pyautogui.press('tab')
    pyautogui.write(senha) #senha
    pyautogui.press('tab')
    time.sleep(5)
    pyautogui.press('enter')
    time.sleep(5)

'''
    Função para preencher os primeiros campos para geração de nota fiscal no site da prefeitura
'''
def prencher_primeiro_campo(CPF,discriminação_dos_Serviços):
    
    pyautogui.PAUSE = 1.5
    
    #preenchendo NFe
    pyautogui.click(412,549) 
    pyautogui.click(482,592)
    time.sleep(2)
    pyautogui.click(482,592)#como muita das vezes o site não reponde corretamente, damos mais um click antes de proseguir
    time.sleep(5) 
    for i in range(3):
        pyautogui.press('tab')
    pyautogui.write(CPF)
    pyautogui.press('tab')
    pyautogui.press('enter')
                
    for i in range(6):
        pyautogui.press('tab')
    pyautogui.press('enter')
    for i in range(2):
        pyautogui.press('tab')
    for i in range(2):
        pyautogui.press('down')
    pyautogui.press('enter')

    for i in range(3):
        pyautogui.press('tab')
    pyautogui.write(discriminação_dos_Serviços)
    pyautogui.press('tab')
    pyautogui.click(975,585)
    time.sleep(2)

'''
    Função para preencher os últimos campos para geração de nota fiscal no site da prefeitura
'''
def prencher_segundo_campo(totalServico,faturamento,alicotaISS,valor_licota,PIS,CONFIRNS,IR,INSS,CSLL,retencoes,msg):
    
    e = faturamento
    g = valor_licota
    #essas duas variaveis servem para selecionar campos que mudam o o fluxo de comandos

    pyautogui.PAUSE = 1.5

    pyautogui.write(totalServico)
    pyautogui.press('tab')
    if e == 'não':
        pyautogui.write(alicotaISS)
        pyautogui.press('tab')
        pyautogui.press('tab')
    elif e == 'sim':
        pyautogui.press('tab')
        pyautogui.press('down')
        pyautogui.press('tab')

    if g == 'valor':
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.write(PIS)
        pyautogui.press('tab')
        pyautogui.write(CONFIRNS)
        pyautogui.press('tab')
        pyautogui.write(IR)
        pyautogui.press('tab')
        pyautogui.write(INSS)
        pyautogui.press('tab')
        pyautogui.write(CSLL)
        pyautogui.press('tab')
        pyautogui.write(retencoes)
        pyautogui.press('tab')

    elif g == 'alíquota' or 'aligota': #as duas opçoes é para evitar erro por causa de digitação
        pyautogui.press('down')
        pyautogui.press('tab')
        pyautogui.write(PIS)
        pyautogui.press('tab')
        pyautogui.write(CONFIRNS)
        pyautogui.press('tab')
        pyautogui.write(IR)
        pyautogui.press('tab')
        pyautogui.write(INSS)
        pyautogui.press('tab')
        pyautogui.write(CSLL)
        pyautogui.press('tab')
        pyautogui.write(retencoes)
        pyautogui.press('tab')

    pyautogui.write(msg)
    #pyautogui.hotkey('ctrl','s') #salva o arquivo

'''
    recebe um arquivo exel e separa as colunas da linha em variaveis e chama outras funçoes que recebem essas
variaveis como entrada.
poaaui um laço de repeticao, para chamar executar as funções em funçao de cada linha da do aquivo ate
que se encontre uma caluna vazia, que é o cpf e é necessario para a NFe, e então ele encerra o processo.
o aequivo possui 14 colunas sendo a primeira o nome do cliente que não é usado para emissão da nota, o
restante das colunas estão na seguinte forma:
[CPF, discriminação dos Serviços prestados,	total dos Serviços,	Faturamento Acima do Limite, aliquotaISS,	
pelo valor ou alíquota,	PIS	, CONFINS ,IR, INSS, CSLL, Outras Retenções, Mensagem Complementar]
'''
def extrair_dados(arquivo):
    book = openpyxl.load_workbook(arquivo)
    page = book.sheetnames[0]
    page = book[page]
    for rows in page.iter_rows(min_row=2):
        
        b = (f'{rows[1].value}')
        c = (f'{rows[2].value}')
        d = (f'{rows[3].value}')
        e = (f'{rows[4].value}')
        f = (f'{rows[5].value}')
        g = (f'{rows[6].value}')
        h = (f'{rows[7].value}')
        i = (f'{rows[8].value}')
        j = (f'{rows[9].value}')
        k = (f'{rows[10].value}')
        l = (f'{rows[11].value}')
        m = (f'{rows[12].value}')
        n = (f'{rows[13].value}')
        if b == 'None':
            break
        else:
            prencher_primeiro_campo(b,c)
            prencher_segundo_campo(d,e,f,g,h,i,j,k,l,m,n)


'''
    Função que invoca as demais para emissão das notas fiscais
'''
def auto(arquivo, senha, cpf, cnpj):
    entrar_na_prefeitura(senha, cpf, cnpj)
    extrair_dados(arquivo)
    
